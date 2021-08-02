import openpyxl

from examples.faker_data import data_samples


def example():
    """
    Создание xlsx файла и запись в него
    """
    # создаю книгу
    book = openpyxl.Workbook()

    # по умолчанию создается с таблицей Sheet
    # print(book.sheetnames)
    book.remove(book.active)

    # создаю таблицы
    # book.active.title = "Коллеги"
    sheet_1 = book.create_sheet("Коллеги")
    sheet_2 = book.create_sheet("Клиенты")
    sheet_3 = book.create_sheet("Черный список", 0)  # таблица будет первой

    for sheet in book.worksheets:  # перебираю таблицы
        for row in data_samples():  # получаю данные
            sheet.append(row)  # записываю данные в строки таблиц

    book.save("test.xlsx")

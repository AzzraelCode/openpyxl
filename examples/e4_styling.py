import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, NamedStyle, Side, Border


def example():
    """
    Оформление таблиц (колонок, строк, ячеек)
    :return:
    """
    filename="test.xlsx"
    book = openpyxl.load_workbook(filename)
    sheet = book.active
    # sheet = book["Коллеги"]

    sheet.column_dimensions["A"].width = 50 # прим. колво символов
    sheet.column_dimensions["B"].width = 20
    # sheet.row_dimensions[1].font = Font(b=True, size=18, color="EE0000") # не будет работать -> Styles can also applied to columns and rows but note that this applies only to cells created (in Excel)

    # стилизация ячеек
    # font = Font(b=True, size=14, color="00DD00")
    # fill1 = PatternFill("darkTrellis") # solid, darkVertical, mediumGray ...
    # fill2 = PatternFill("solid", fgColor="FFFF99")
    # sheet['A3'].font = font
    # sheet['A3'].fill = fill1
    # sheet['B3'].font = font
    # sheet['B3'].fill = fill2

    # встроенные стили https://openpyxl.readthedocs.io/en/stable/styles.html?highlight=styles
    # sheet['A2'].style = "Good"
    # sheet['B2'].style = "Bad"

    # создание именованного стиля
    # azzcode_style = NamedStyle(name="azzcode_style")
    # azzcode_style.font = Font(b=True, size=14, color="DD0000")
    # azzcode_style.fill = PatternFill("solid", fgColor="FFFF99")
    # side = Side(style='medium', color="00EEDD") # 'dashDot','dashDotDot', 'dashed','dotted',; 'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',; 'mediumDashed', 'slantDashDot', 'thick', 'thin'
    # azzcode_style.border = Border(bottom=side)
    #
    # sheet['A1'].style = azzcode_style
    # sheet['B1'].style = azzcode_style
    # sheet['C1'].style = azzcode_style

    book.save(filename)